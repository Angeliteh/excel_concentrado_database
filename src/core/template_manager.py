"""
📋 TEMPLATE MANAGER - Gestor Especializado de Plantillas
=======================================================

Módulo especializado únicamente en gestionar plantillas Excel.
Responsabilidad única: validación, carga y gestión de plantillas.

CARACTERÍSTICAS:
✅ Solo gestión de plantillas (responsabilidad única)
✅ Validación de plantillas
✅ Información de plantillas
✅ Sin lógica de datos o escritura
✅ Extensible y testeable
"""

import os
from pathlib import Path
from openpyxl import load_workbook
from typing import Dict, List, Any, Optional
from ..config.settings import get_config_actual


class TemplateManager:
    """
    Gestor especializado de plantillas con responsabilidad única.
    
    Solo se encarga de gestionar plantillas Excel,
    sin lógica de datos o escritura.
    """
    
    def __init__(self):
        """Inicializar gestor de plantillas."""
        self.config_actual = get_config_actual()
        self.plantillas_cache = {}
        print("📋 TemplateManager inicializado")
    
    def validar_plantilla(self, plantilla_path: str) -> Dict[str, Any]:
        """
        Validar que una plantilla existe y es válida.
        
        Args:
            plantilla_path: Ruta de la plantilla
            
        Returns:
            dict: Resultado de validación
        """
        try:
            # Verificar que el archivo existe
            if not os.path.exists(plantilla_path):
                return {
                    'valida': False,
                    'mensaje': f'Plantilla no encontrada: {plantilla_path}',
                    'error': 'FileNotFoundError'
                }
            
            # Intentar cargar la plantilla
            workbook = load_workbook(plantilla_path)
            hojas = workbook.sheetnames
            
            # Validaciones básicas
            if len(hojas) == 0:
                workbook.close()
                return {
                    'valida': False,
                    'mensaje': 'Plantilla sin hojas',
                    'error': 'NoSheetsError'
                }
            
            # Obtener información básica
            hoja_principal = workbook[hojas[0]]
            dimensiones = {
                'max_row': hoja_principal.max_row,
                'max_column': hoja_principal.max_column
            }
            
            workbook.close()
            
            return {
                'valida': True,
                'mensaje': 'Plantilla válida',
                'hojas': hojas,
                'hoja_principal': hojas[0],
                'dimensiones': dimensiones
            }
            
        except Exception as e:
            return {
                'valida': False,
                'mensaje': f'Error validando plantilla: {str(e)}',
                'error': type(e).__name__
            }
    
    def obtener_info_plantilla(self, plantilla_path: str) -> Dict[str, Any]:
        """
        Obtener información detallada de una plantilla.
        
        Args:
            plantilla_path: Ruta de la plantilla
            
        Returns:
            dict: Información detallada de la plantilla
        """
        try:
            # Usar cache si está disponible
            if plantilla_path in self.plantillas_cache:
                print(f"📋 Usando info de plantilla desde cache: {plantilla_path}")
                return self.plantillas_cache[plantilla_path]
            
            # Validar primero
            validacion = self.validar_plantilla(plantilla_path)
            if not validacion['valida']:
                return validacion
            
            # Obtener información detallada
            workbook = load_workbook(plantilla_path)
            
            info = {
                'valida': True,
                'archivo': {
                    'path': plantilla_path,
                    'nombre': os.path.basename(plantilla_path),
                    'tamaño': os.path.getsize(plantilla_path),
                    'extension': Path(plantilla_path).suffix
                },
                'hojas': [],
                'celdas_combinadas_total': 0,
                'propiedades': {}
            }
            
            # Información de cada hoja
            for nombre_hoja in workbook.sheetnames:
                hoja = workbook[nombre_hoja]
                
                info_hoja = {
                    'nombre': nombre_hoja,
                    'dimensiones': {
                        'max_row': hoja.max_row,
                        'max_column': hoja.max_column,
                        'celdas_usadas': hoja.max_row * hoja.max_column
                    },
                    'celdas_combinadas': len(list(hoja.merged_cells.ranges)),
                    'es_principal': nombre_hoja == workbook.sheetnames[0]
                }
                
                info['hojas'].append(info_hoja)
                info['celdas_combinadas_total'] += info_hoja['celdas_combinadas']
            
            # Propiedades del workbook
            if hasattr(workbook, 'properties'):
                props = workbook.properties
                info['propiedades'] = {
                    'titulo': getattr(props, 'title', ''),
                    'autor': getattr(props, 'creator', ''),
                    'descripcion': getattr(props, 'description', ''),
                    'creado': getattr(props, 'created', None),
                    'modificado': getattr(props, 'modified', None)
                }
            
            workbook.close()
            
            # Guardar en cache
            self.plantillas_cache[plantilla_path] = info
            
            print(f"📋 Info de plantilla obtenida: {info['archivo']['nombre']}")
            return info
            
        except Exception as e:
            return {
                'valida': False,
                'mensaje': f'Error obteniendo info de plantilla: {str(e)}',
                'error': type(e).__name__
            }
    
    def buscar_plantillas_en_directorio(self, directorio: str, 
                                       patron: str = "*.xlsx") -> List[Dict[str, Any]]:
        """
        Buscar plantillas en un directorio.
        
        Args:
            directorio: Directorio donde buscar
            patron: Patrón de archivos a buscar
            
        Returns:
            list: Lista de plantillas encontradas con su información
        """
        try:
            if not os.path.exists(directorio):
                print(f"❌ Directorio no encontrado: {directorio}")
                return []
            
            plantillas_encontradas = []
            directorio_path = Path(directorio)
            
            # Buscar archivos que coincidan con el patrón
            for archivo in directorio_path.glob(patron):
                if archivo.is_file():
                    # Validar cada plantilla encontrada
                    validacion = self.validar_plantilla(str(archivo))
                    
                    plantilla_info = {
                        'path': str(archivo),
                        'nombre': archivo.name,
                        'valida': validacion['valida'],
                        'mensaje': validacion['mensaje']
                    }
                    
                    if validacion['valida']:
                        plantilla_info.update({
                            'hojas': validacion['hojas'],
                            'dimensiones': validacion['dimensiones']
                        })
                    
                    plantillas_encontradas.append(plantilla_info)
            
            print(f"📋 Plantillas encontradas en {directorio}: {len(plantillas_encontradas)}")
            return plantillas_encontradas
            
        except Exception as e:
            print(f"❌ Error buscando plantillas: {e}")
            return []
    
    def obtener_plantilla_por_defecto(self) -> Optional[str]:
        """
        Obtener la ruta de la plantilla por defecto según configuración.
        
        Returns:
            str: Ruta de la plantilla por defecto o None si no se encuentra
        """
        try:
            # Buscar plantilla por defecto en configuración
            plantilla_default = "plantilla_base.xlsx"
            
            # Buscar en directorio actual
            if os.path.exists(plantilla_default):
                validacion = self.validar_plantilla(plantilla_default)
                if validacion['valida']:
                    print(f"📋 Plantilla por defecto encontrada: {plantilla_default}")
                    return plantilla_default
            
            # Buscar en directorios comunes
            directorios_busqueda = [
                ".",
                "templates",
                "plantillas",
                "assets"
            ]
            
            for directorio in directorios_busqueda:
                plantilla_path = os.path.join(directorio, plantilla_default)
                if os.path.exists(plantilla_path):
                    validacion = self.validar_plantilla(plantilla_path)
                    if validacion['valida']:
                        print(f"📋 Plantilla por defecto encontrada en {directorio}: {plantilla_path}")
                        return plantilla_path
            
            print("⚠️ Plantilla por defecto no encontrada")
            return None
            
        except Exception as e:
            print(f"❌ Error obteniendo plantilla por defecto: {e}")
            return None
    
    def limpiar_cache(self):
        """Limpiar cache de plantillas."""
        self.plantillas_cache.clear()
        print("🧹 Cache de plantillas limpiado")
    
    def verificar_compatibilidad_plantilla(self, plantilla_path: str, 
                                         modo: str = None) -> Dict[str, Any]:
        """
        Verificar compatibilidad de plantilla con el modo actual.
        
        Args:
            plantilla_path: Ruta de la plantilla
            modo: Modo específico a verificar (opcional)
            
        Returns:
            dict: Resultado de verificación de compatibilidad
        """
        try:
            if modo is None:
                modo = self.config_actual.get('MODO', 'ESCUELAS')
            
            # Obtener info de la plantilla
            info = self.obtener_info_plantilla(plantilla_path)
            if not info['valida']:
                return info
            
            # Verificaciones específicas por modo
            compatibilidad = {
                'compatible': True,
                'modo': modo,
                'advertencias': [],
                'errores': []
            }
            
            if modo == 'ESCUELAS':
                # Verificar que tenga al menos las dimensiones mínimas esperadas
                hoja_principal = info['hojas'][0] if info['hojas'] else None
                if hoja_principal:
                    if hoja_principal['dimensiones']['max_column'] < 26:  # Menos de Z
                        compatibilidad['advertencias'].append(
                            f"Plantilla tiene solo {hoja_principal['dimensiones']['max_column']} columnas, "
                            "se esperan al menos 26 (A-Z)"
                        )
                    
                    if hoja_principal['dimensiones']['max_row'] < 15:
                        compatibilidad['advertencias'].append(
                            f"Plantilla tiene solo {hoja_principal['dimensiones']['max_row']} filas, "
                            "se esperan al menos 15"
                        )
            
            elif modo == 'ZONAS':
                # Verificaciones específicas para modo ZONAS
                pass
            
            # Determinar compatibilidad final
            if compatibilidad['errores']:
                compatibilidad['compatible'] = False
            
            return compatibilidad
            
        except Exception as e:
            return {
                'compatible': False,
                'mensaje': f'Error verificando compatibilidad: {str(e)}',
                'error': type(e).__name__
            }
