"""
📝 EXCEL WRITER - Escritor Especializado de Excel
================================================

Módulo especializado únicamente en escribir datos a archivos Excel.
Responsabilidad única: operaciones de escritura en Excel.

CARACTERÍSTICAS:
✅ Solo escritura de Excel (responsabilidad única)
✅ Manejo de celdas combinadas
✅ Configuración centralizada
✅ Sin lógica de mapeo o gestión de plantillas
✅ Extensible y testeable
"""

from openpyxl import load_workbook
from typing import Dict, Any, List, Tuple
from ..config.ui_config import get_ui_config


class ExcelWriter:
    """
    Escritor especializado de Excel con responsabilidad única.
    
    Solo se encarga de escribir datos en archivos Excel,
    sin lógica de mapeo o gestión de plantillas.
    """
    
    def __init__(self):
        """Inicializar escritor de Excel."""
        self.ui_config = get_ui_config()
        print("📝 ExcelWriter inicializado")
    
    def escribir_datos_en_hoja(self, archivo_path: str, hoja_nombre: str, 
                              datos_mapeados: List[Tuple[int, int, Any]], 
                              preservar_combinadas: bool = True) -> bool:
        """
        Escribir datos mapeados en una hoja de Excel.
        
        Args:
            archivo_path: Ruta del archivo Excel
            hoja_nombre: Nombre de la hoja
            datos_mapeados: Lista de tuplas (fila, columna, valor)
            preservar_combinadas: Si preservar celdas combinadas
            
        Returns:
            bool: True si la escritura fue exitosa
        """
        try:
            print(f"📝 Escribiendo datos en {archivo_path}, hoja '{hoja_nombre}'")
            
            # Cargar workbook
            workbook = load_workbook(archivo_path)
            
            if hoja_nombre not in workbook.sheetnames:
                print(f"❌ Hoja '{hoja_nombre}' no encontrada")
                return False
            
            hoja = workbook[hoja_nombre]
            
            # Obtener mapa de celdas combinadas si es necesario
            mapa_combinadas = {}
            if preservar_combinadas:
                mapa_combinadas = self._crear_mapa_combinadas(hoja.merged_cells.ranges)
            
            # Escribir cada dato
            datos_escritos = 0
            for fila, columna, valor in datos_mapeados:
                if self._escribir_valor_en_celda(hoja, fila, columna, valor, mapa_combinadas):
                    datos_escritos += 1
            
            # Guardar archivo
            workbook.save(archivo_path)
            workbook.close()
            
            print(f"✅ Escritura completada: {datos_escritos} valores escritos")
            return True
            
        except Exception as e:
            print(f"❌ Error escribiendo en Excel: {e}")
            return False
    
    def _escribir_valor_en_celda(self, hoja, fila: int, columna: int, valor: Any, 
                                mapa_combinadas: Dict) -> bool:
        """
        Escribir un valor en una celda específica.
        
        Args:
            hoja: Hoja de Excel (openpyxl)
            fila: Número de fila (1-based)
            columna: Número de columna (1-based)
            valor: Valor a escribir
            mapa_combinadas: Mapa de celdas combinadas
            
        Returns:
            bool: True si se escribió correctamente
        """
        try:
            # Verificar si la celda está en un rango combinado
            if (fila, columna) in mapa_combinadas:
                rango_combinado = mapa_combinadas[(fila, columna)]
                
                # Solo escribir en la celda principal del rango combinado
                if fila == rango_combinado.min_row and columna == rango_combinado.min_col:
                    celda = hoja.cell(row=fila, column=columna)
                    celda.value = valor
                    print(f"   📝 Escrito en celda principal combinada {celda.coordinate}: {valor}")
                    return True
                else:
                    # Celda secundaria de rango combinado - no escribir
                    print(f"   ⏭️ Saltando celda secundaria combinada ({fila},{columna})")
                    return False
            else:
                # Celda normal - escribir directamente
                celda = hoja.cell(row=fila, column=columna)
                celda.value = valor
                print(f"   📝 Escrito en celda normal {celda.coordinate}: {valor}")
                return True
                
        except Exception as e:
            print(f"   ❌ Error escribiendo en celda ({fila},{columna}): {e}")
            return False
    
    def _crear_mapa_combinadas(self, rangos_combinados) -> Dict[Tuple[int, int], Any]:
        """
        Crear mapa de celdas combinadas para referencia rápida.
        
        Args:
            rangos_combinados: Rangos combinados de openpyxl
            
        Returns:
            dict: Mapa de (fila, columna) -> rango_combinado
        """
        mapa_combinadas = {}
        
        for rango_combinado in rangos_combinados:
            for fila in range(rango_combinado.min_row, rango_combinado.max_row + 1):
                for col in range(rango_combinado.min_col, rango_combinado.max_col + 1):
                    mapa_combinadas[(fila, col)] = rango_combinado
        
        print(f"🔗 Mapa de celdas combinadas creado: {len(rangos_combinados)} rangos")
        return mapa_combinadas
    
    def validar_archivo_escribible(self, archivo_path: str) -> Dict[str, Any]:
        """
        Validar que un archivo Excel es escribible.
        
        Args:
            archivo_path: Ruta del archivo
            
        Returns:
            dict: Resultado de validación
        """
        try:
            # Intentar abrir y cerrar el archivo
            workbook = load_workbook(archivo_path)
            hojas = workbook.sheetnames
            workbook.close()
            
            return {
                'escribible': True,
                'hojas_disponibles': hojas,
                'mensaje': 'Archivo válido y escribible'
            }
            
        except FileNotFoundError:
            return {
                'escribible': False,
                'mensaje': 'Archivo no encontrado',
                'error': 'FileNotFoundError'
            }
        except PermissionError:
            return {
                'escribible': False,
                'mensaje': 'Sin permisos de escritura',
                'error': 'PermissionError'
            }
        except Exception as e:
            return {
                'escribible': False,
                'mensaje': f'Error validando archivo: {str(e)}',
                'error': type(e).__name__
            }
    
    def crear_backup(self, archivo_path: str) -> str:
        """
        Crear backup de un archivo antes de modificarlo.
        
        Args:
            archivo_path: Ruta del archivo original
            
        Returns:
            str: Ruta del archivo de backup creado
        """
        import shutil
        from datetime import datetime
        
        try:
            # Generar nombre de backup con timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_path = archivo_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
            
            # Copiar archivo
            shutil.copy2(archivo_path, backup_path)
            
            print(f"💾 Backup creado: {backup_path}")
            return backup_path
            
        except Exception as e:
            print(f"❌ Error creando backup: {e}")
            raise e
    
    def obtener_info_hoja(self, archivo_path: str, hoja_nombre: str) -> Dict[str, Any]:
        """
        Obtener información de una hoja específica.
        
        Args:
            archivo_path: Ruta del archivo
            hoja_nombre: Nombre de la hoja
            
        Returns:
            dict: Información de la hoja
        """
        try:
            workbook = load_workbook(archivo_path)
            
            if hoja_nombre not in workbook.sheetnames:
                return {
                    'existe': False,
                    'mensaje': f"Hoja '{hoja_nombre}' no encontrada"
                }
            
            hoja = workbook[hoja_nombre]
            
            info = {
                'existe': True,
                'dimensiones': {
                    'max_row': hoja.max_row,
                    'max_column': hoja.max_column
                },
                'celdas_combinadas': len(list(hoja.merged_cells.ranges)),
                'titulo': hoja.title
            }
            
            workbook.close()
            return info
            
        except Exception as e:
            return {
                'existe': False,
                'mensaje': f'Error obteniendo info: {str(e)}',
                'error': type(e).__name__
            }
