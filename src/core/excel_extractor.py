"""
📊 EXCEL EXTRACTOR - Extracción Pura de Datos
============================================

Módulo especializado ÚNICAMENTE en extraer datos de archivos Excel.
Responsabilidad única: leer datos sin transformarlos.

PRINCIPIOS:
✅ Una sola responsabilidad: extraer datos
✅ Sin transformaciones: datos tal como están en Excel
✅ Sin lógica de negocio: solo lectura
✅ Reutilizable: funciona con cualquier hoja/rango
✅ Configurable: usa esquemas dinámicos
"""

import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional, Any
from ..utils.excel_utils import cargar_excel


class ExcelExtractor:
    """
    Extractor puro de datos de Excel.
    
    Responsabilidad única: leer datos de archivos Excel sin transformarlos.
    """
    
    def __init__(self):
        """Inicializar extractor."""
        print("📊 ExcelExtractor inicializado")
    
    def extraer_hoja_simple(self, archivo_path: str, hoja_nombre: str, rango: str) -> pd.DataFrame:
        """
        Extraer datos de una hoja específica en un rango dado.
        
        Args:
            archivo_path: Ruta del archivo Excel
            hoja_nombre: Nombre de la hoja
            rango: Rango en formato "A5:Z17"
            
        Returns:
            DataFrame con datos tal como están en Excel
        """
        print(f"📋 Extrayendo hoja '{hoja_nombre}' rango '{rango}' de {archivo_path}")
        
        # Cargar hoja
        hoja = cargar_excel(archivo_path, hoja_nombre)
        
        # Parsear rango
        rango_coords = self._parsear_rango(rango)
        
        # Extraer datos sin transformar
        datos_raw = self._extraer_datos_raw(hoja, rango_coords)
        
        # Crear DataFrame
        df = pd.DataFrame(datos_raw)
        
        print(f"✅ Datos extraídos: {df.shape}")
        return df
    
    def extraer_multiples_hojas(self, archivo_path: str, config_hojas: Dict[str, Dict]) -> Dict[str, pd.DataFrame]:
        """
        Extraer datos de múltiples hojas del mismo archivo.
        
        Args:
            archivo_path: Ruta del archivo Excel
            config_hojas: Configuración por hoja
                {
                    "ESC1": {"hoja": "ESC1", "rango": "A3:Z15"},
                    "ESC2": {"hoja": "ESC2", "rango": "A5:Z17"}
                }
                
        Returns:
            Diccionario con DataFrames por hoja
        """
        print(f"📊 Extrayendo múltiples hojas de {archivo_path}")
        
        resultados = {}
        
        for nombre_config, config in config_hojas.items():
            try:
                hoja_nombre = config['hoja']
                rango = config['rango']
                
                datos = self.extraer_hoja_simple(archivo_path, hoja_nombre, rango)
                resultados[nombre_config] = datos
                
                print(f"✅ {nombre_config}: {datos.shape}")
                
            except Exception as e:
                print(f"❌ Error extrayendo {nombre_config}: {e}")
                resultados[nombre_config] = None
        
        return resultados
    
    def detectar_celdas_combinadas(self, archivo_path: str, hoja_nombre: str) -> List[Tuple]:
        """
        Detectar celdas combinadas en una hoja.
        
        Args:
            archivo_path: Ruta del archivo Excel
            hoja_nombre: Nombre de la hoja
            
        Returns:
            Lista de rangos combinados como tuplas (min_row, min_col, max_row, max_col)
        """
        print(f"🔍 Detectando celdas combinadas en '{hoja_nombre}'")
        
        # Cargar workbook para acceder a merged_cells
        workbook = load_workbook(archivo_path, data_only=True)
        hoja = workbook[hoja_nombre]
        
        # Obtener rangos combinados
        rangos_combinados = []
        for rango in hoja.merged_cells.ranges:
            coords = (rango.min_row, rango.min_col, rango.max_row, rango.max_col)
            rangos_combinados.append(coords)
        
        print(f"✅ Detectadas {len(rangos_combinados)} celdas combinadas")
        return rangos_combinados
    
    def extraer_con_metadatos(self, archivo_path: str, hoja_nombre: str, rango: str) -> Dict[str, Any]:
        """
        Extraer datos junto con metadatos (celdas combinadas, tipos, etc.).
        
        Args:
            archivo_path: Ruta del archivo Excel
            hoja_nombre: Nombre de la hoja
            rango: Rango en formato "A5:Z17"
            
        Returns:
            Diccionario con datos y metadatos
        """
        print(f"📊 Extrayendo con metadatos: '{hoja_nombre}' rango '{rango}'")
        
        # Extraer datos principales
        datos = self.extraer_hoja_simple(archivo_path, hoja_nombre, rango)
        
        # Extraer metadatos
        celdas_combinadas = self.detectar_celdas_combinadas(archivo_path, hoja_nombre)
        
        # Filtrar celdas combinadas que están en el rango
        rango_coords = self._parsear_rango(rango)
        celdas_en_rango = self._filtrar_celdas_en_rango(celdas_combinadas, rango_coords)
        
        resultado = {
            'datos': datos,
            'celdas_combinadas': celdas_en_rango,
            'rango_original': rango,
            'hoja_nombre': hoja_nombre,
            'archivo_path': archivo_path,
            'dimensiones': datos.shape
        }
        
        print(f"✅ Extracción con metadatos completada")
        return resultado
    
    def _parsear_rango(self, rango: str) -> Dict[str, int]:
        """
        Parsear rango de Excel a coordenadas.
        
        Args:
            rango: Rango en formato "A5:Z17"
            
        Returns:
            Diccionario con coordenadas
        """
        partes = rango.split(':')
        inicio = partes[0]  # "A5"
        fin = partes[1]     # "Z17"
        
        # Extraer números de fila
        min_row = int(''.join(filter(str.isdigit, inicio)))
        max_row = int(''.join(filter(str.isdigit, fin)))
        
        # Extraer letras de columna y convertir a números
        inicio_col = inicio.rstrip('0123456789')  # "A"
        fin_col = fin.rstrip('0123456789')        # "Z"
        
        min_col = self._letra_a_numero(inicio_col)
        max_col = self._letra_a_numero(fin_col)
        
        return {
            'min_row': min_row,
            'max_row': max_row,
            'min_col': min_col,
            'max_col': max_col
        }
    
    def _letra_a_numero(self, letra: str) -> int:
        """
        Convertir letra de columna a número (A=1, B=2, ..., Z=26).
        
        Args:
            letra: Letra de columna (ej: "A", "Z", "AA")
            
        Returns:
            Número de columna
        """
        numero = 0
        for char in letra:
            numero = numero * 26 + (ord(char.upper()) - ord('A') + 1)
        return numero
    
    def _extraer_datos_raw(self, hoja, rango_coords: Dict[str, int]) -> List[List]:
        """
        Extraer datos raw de la hoja en el rango especificado.
        
        Args:
            hoja: Hoja de Excel (openpyxl)
            rango_coords: Coordenadas del rango
            
        Returns:
            Lista de listas con datos raw
        """
        datos = []
        
        for fila in range(rango_coords['min_row'], rango_coords['max_row'] + 1):
            fila_datos = []
            for col in range(rango_coords['min_col'], rango_coords['max_col'] + 1):
                celda = hoja.cell(row=fila, column=col)
                valor = celda.value if celda.value is not None else ""
                fila_datos.append(valor)
            datos.append(fila_datos)
        
        return datos
    
    def _filtrar_celdas_en_rango(self, celdas_combinadas: List[Tuple], rango_coords: Dict[str, int]) -> List[Tuple]:
        """
        Filtrar celdas combinadas que están dentro del rango especificado.
        
        Args:
            celdas_combinadas: Lista de tuplas con rangos combinados
            rango_coords: Coordenadas del rango de interés
            
        Returns:
            Lista filtrada de celdas combinadas
        """
        celdas_filtradas = []
        
        for celda in celdas_combinadas:
            min_row, min_col, max_row, max_col = celda
            
            # Verificar si la celda combinada está dentro del rango
            if (min_row >= rango_coords['min_row'] and max_row <= rango_coords['max_row'] and
                min_col >= rango_coords['min_col'] and max_col <= rango_coords['max_col']):
                
                # Ajustar coordenadas relativas al rango
                celda_relativa = (
                    min_row - rango_coords['min_row'],
                    min_col - rango_coords['min_col'],
                    max_row - rango_coords['min_row'],
                    max_col - rango_coords['min_col']
                )
                celdas_filtradas.append(celda_relativa)
        
        return celdas_filtradas
    
    def validar_archivo(self, archivo_path: str) -> bool:
        """
        Validar que el archivo Excel existe y es accesible.
        
        Args:
            archivo_path: Ruta del archivo
            
        Returns:
            True si el archivo es válido
        """
        try:
            workbook = load_workbook(archivo_path, data_only=True)
            workbook.close()
            return True
        except Exception as e:
            print(f"❌ Error validando archivo {archivo_path}: {e}")
            return False
    
    def listar_hojas(self, archivo_path: str) -> List[str]:
        """
        Listar todas las hojas disponibles en el archivo.
        
        Args:
            archivo_path: Ruta del archivo Excel
            
        Returns:
            Lista de nombres de hojas
        """
        try:
            workbook = load_workbook(archivo_path, data_only=True)
            hojas = workbook.sheetnames
            workbook.close()
            print(f"📋 Hojas disponibles: {hojas}")
            return hojas
        except Exception as e:
            print(f"❌ Error listando hojas: {e}")
            return []
