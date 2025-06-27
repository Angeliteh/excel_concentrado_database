import os
import win32com.client
import pythoncom
from openpyxl import load_workbook
import pandas as pd
import time
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def extraer_tabla_y_limpiar(hoja, rango):
    """
    Extrae y limpia datos de un rango de Excel
    
    Args:
        hoja: Hoja de Excel
        rango: Tupla (fila_inicio, col_inicio, fila_fin, col_fin)
    """
    min_row, min_col, max_row, max_col = rango
    
    # Manejar celdas combinadas
    rangos_combinados = list(hoja.merged_cells.ranges)
    for rango_combinado in rangos_combinados:
        if (rango_combinado.min_row >= min_row and 
            rango_combinado.max_row <= max_row and
            rango_combinado.min_col >= min_col and 
            rango_combinado.max_col <= max_col):
            
            valor = hoja.cell(rango_combinado.min_row, rango_combinado.min_col).value
            hoja.unmerge_cells(rango_combinado.coord)
            
            for fila in range(rango_combinado.min_row, rango_combinado.max_row + 1):
                for col in range(rango_combinado.min_col, rango_combinado.max_col + 1):
                    hoja.cell(row=fila, column=col, value=valor)

    # Extraer datos
    tabla = []
    for fila in range(min_row, max_row + 1):
        fila_datos = []
        for col in range(min_col, max_col + 1):
            valor = hoja.cell(row=fila, column=col).value
            fila_datos.append(valor)
        tabla.append(fila_datos)
    
    return tabla

def cargar_excel(archivo, hoja_nombre):
    """
    Carga un archivo Excel y retorna la hoja especificada
    """
    try:
        wb = load_workbook(archivo, data_only=True)
        return wb[hoja_nombre]
    except Exception as e:
        raise Exception(f"Error al cargar el archivo Excel: {str(e)}")

def inyectar_datos_en_plantilla(df, hoja, rango_inyeccion):
    """
    Inyecta datos en la plantilla respetando las celdas combinadas
    """
    min_row, min_col, max_row, max_col = rango_inyeccion
    
    # Crear un mapa de celdas combinadas
    merged_ranges = {}
    for merged_range in hoja.merged_cells.ranges:
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_ranges[(row, col)] = merged_range

    # Inyectar datos
    for i, row in enumerate(df.values):
        for j, value in enumerate(row):
            current_row = min_row + i
            current_col = min_col + j
            
            # Obtener la celda actual
            cell = hoja.cell(row=current_row, column=current_col)
            
            # Si es una celda combinada, obtener la celda principal
            if isinstance(cell, MergedCell):
                merged_range = merged_ranges.get((current_row, current_col))
                if merged_range:
                    # Usar la celda superior izquierda del rango combinado
                    cell = hoja.cell(row=merged_range.min_row, column=merged_range.min_col)
            
            # Establecer el valor
            if not isinstance(cell, MergedCell):
                cell.value = value

def inyectar_formulas_totales_y_subtotales(archivo_salida, hoja_nombre, fila_inicial, fila_final):
    """
    Inyecta las fórmulas exactamente como en el script antiguo
    """
    wb = load_workbook(archivo_salida)
    hoja = wb[hoja_nombre]

    for fila in range(fila_inicial, fila_final + 1):
        hoja[f"T{fila}"] = f"=H{fila}+J{fila}+L{fila}+N{fila}+P{fila}+R{fila}"
        hoja[f"V{fila}"] = f"=I{fila}+K{fila}+M{fila}+O{fila}+Q{fila}+S{fila}"
        hoja[f"X{fila}"] = f"=T{fila}+V{fila}"

    wb.save(archivo_salida)

def convertir_formulas_a_valores(archivo):
    """
    Convierte todas las fórmulas a valores en el archivo
    """
    import win32com.client
    import pythoncom

    try:
        # Inicializar COM
        pythoncom.CoInitialize()

        # Intentar crear la aplicación Excel
        try:
            excel = win32com.client.Dispatch("Excel.Application")
        except Exception as e:
            print(f"Error al inicializar Excel COM: {e}")
            print("Saltando conversión de fórmulas a valores...")
            return

        excel.Visible = False
        excel.DisplayAlerts = False

        # Convertir a ruta absoluta de forma segura
        ruta_absoluta = os.path.abspath(archivo)
        print(f"Abriendo archivo para conversión: {ruta_absoluta}")

        try:
            wb = excel.Workbooks.Open(ruta_absoluta)
            print("Archivo abierto exitosamente")

            for ws in wb.Worksheets:
                print(f"Procesando hoja: {ws.Name}")
                # Convertir fórmulas a valores
                if ws.UsedRange is not None:
                    ws.UsedRange.Value = ws.UsedRange.Value

            wb.Save()
            print("Archivo guardado exitosamente")
            wb.Close()

        except Exception as e:
            print(f"Error al procesar el archivo: {e}")
            try:
                wb.Close(SaveChanges=False)
            except:
                pass

    except Exception as e:
        print(f"Error general en conversión de fórmulas: {e}")

    finally:
        try:
            excel.Quit()
        except:
            pass
        try:
            pythoncom.CoUninitialize()
        except:
            pass
